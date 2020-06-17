from trello import TrelloAPI as Trello
from openpyxl import load_workbook


def main():
    api_key = "YOUR_API_KEY_HERE"
    token = "YOUR_TOKEN_HERE"
    board_id = "BOARD_ID_HERE"
    list_name = "LIST_NAME_HERE"
    xlsx_file_path = "data.xlsx"
    label_name = "LABEL_NAME_HERE"

    MyTrello = Trello(api_key,token)
    B = MyTrello.GetBoardById(board_id)
    TodoList = B.GetListByName(list_name)
    wb = load_workbook(xlsx_file_path)
    sheet = wb.active
    max_row = sheet.max_row
    label = B.GetLabelByName(label_name)
    for i in range(2, max_row+1):
        # iterate over all columns
        print(i)
        question = sheet.cell(row=i, column=1).value
        Card = MyTrello.new('Card', question, TodoList)
        comment = sheet.cell(row=i, column=2).value
        # Test if comment is empty
        if comment is not None:
            Card.SetDesc(comment)
        priority = sheet.cell(row=i, column=3).value
        difficulty = sheet.cell(row=i, column=4).value
        Card.AddLabel(label)
        

if __name__ == '__main__':
    main()
